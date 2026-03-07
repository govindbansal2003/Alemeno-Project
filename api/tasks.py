import os
import logging
from datetime import datetime

from celery import shared_task
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


@shared_task
def ingest_customer_data():
    """Ingest customer data from customer_data.xlsx into the database."""
    from api.models import Customer

    file_path = os.path.join(BASE_DIR, 'files', 'customer_data.xlsx')
    if not os.path.exists(file_path):
        logger.error(f"Customer data file not found: {file_path}")
        return "File not found"

    wb = load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    count = 0

    for row in rows:
        if not row or row[0] is None:
            continue

        customer_id = int(row[0])
        first_name = str(row[1]).strip() if row[1] else ''
        last_name = str(row[2]).strip() if row[2] else ''
        phone_number = str(int(row[3])) if row[3] else '0'
        monthly_salary = int(row[4]) if row[4] else 0
        approved_limit = int(row[5]) if row[5] else 0
        current_debt = float(row[6]) if row[6] else 0

        Customer.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'phone_number': phone_number,
                'monthly_salary': monthly_salary,
                'approved_limit': approved_limit,
                'current_debt': current_debt,
            }
        )
        count += 1

    wb.close()
    logger.info(f"Ingested {count} customers")
    return f"Ingested {count} customers"


@shared_task
def ingest_loan_data():
    """Ingest loan data from loan_data.xlsx into the database."""
    from api.models import Customer, Loan

    file_path = os.path.join(BASE_DIR, 'files', 'loan_data.xlsx')
    if not os.path.exists(file_path):
        logger.error(f"Loan data file not found: {file_path}")
        return "File not found"

    wb = load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    count = 0

    for row in rows:
        if not row or row[0] is None:
            continue

        customer_id = int(row[0])
        loan_id = int(row[1])
        loan_amount = float(row[2]) if row[2] else 0
        tenure = int(row[3]) if row[3] else 0
        interest_rate = float(row[4]) if row[4] else 0
        monthly_repayment = float(row[5]) if row[5] else 0
        emis_paid_on_time = int(row[6]) if row[6] else 0

        # Parse dates
        start_date = None
        end_date = None
        if row[7]:
            if isinstance(row[7], datetime):
                start_date = row[7].date()
            else:
                try:
                    start_date = datetime.strptime(str(row[7]), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
        if row[8]:
            if isinstance(row[8], datetime):
                end_date = row[8].date()
            else:
                try:
                    end_date = datetime.strptime(str(row[8]), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass

        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            logger.warning(f"Customer {customer_id} not found, skipping loan {loan_id}")
            continue

        Loan.objects.update_or_create(
            loan_id=loan_id,
            defaults={
                'customer': customer,
                'loan_amount': loan_amount,
                'tenure': tenure,
                'interest_rate': interest_rate,
                'monthly_repayment': monthly_repayment,
                'emis_paid_on_time': emis_paid_on_time,
                'start_date': start_date,
                'end_date': end_date,
            }
        )
        count += 1

    wb.close()
    logger.info(f"Ingested {count} loans")
    return f"Ingested {count} loans"


@shared_task
def ingest_all_data():
    """Ingest both customer and loan data."""
    customer_result = ingest_customer_data()
    loan_result = ingest_loan_data()
    return f"{customer_result}; {loan_result}"
